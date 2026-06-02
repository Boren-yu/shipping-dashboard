import openpyxl, re, json, math, statistics
from collections import defaultdict

EXCEL = '../data/国际海运费用招标信息汇总.xlsx'

PERIOD_MAP = {
    '25.Q2':'Q2 2025','25.7-8月':'25年7-8月','25.9月':'25年9月',
    '25.10月':'25年10月','25.11月':'25年11月','25.12月':'25年12月',
    '26.1月':'26年1月','26.2月':'26年2月','26.3月':'26年3月',
    '26.3.30-4.12':'26年4月上','26.4.13-4.26':'26年4月下',
    '26.5月':'26年5月','26.6月':'26年6月',
}
ALL_PERIODS = ['Q2 2025','25年7-8月','25年9月','25年10月','25年11月','25年12月',
               '26年1月','26年2月','26年3月','26年4月上','26年4月下','26年5月','26年6月']
CURR_PERIOD = '26年5月'
CURR_IDX = ALL_PERIODS.index(CURR_PERIOD)

PORT_NORMS = {
    'sha':'SHA','sai':'SAI','dalian':'Dalian','tianjin':'Tianjin',
    'incheon':'Incheon','bangkok':'Bangkok','bkk':'Bangkok','dubai':'Dubai',
    'laem':'Laem Chabang','laemchabang':'Laem Chabang',
    'nhava':'Nhava Sheva','nhavasheva':'Nhava Sheva',
    'izmit':'Izmit','istanbul':'Izmit','manzanillo':'Manzanillo',
    'hou':'HOU','houston':'HOU','detroit':'Detroit','savanna':'Savanna','savannah':'Savanna',
    'hamburg':'Hamburg','koper':'Koper','shanghai':'SHA',
}

def normalize_port(s):
    s = s.strip().rstrip(',')
    key = re.sub(r'[\s\-,]','',s.lower())
    return PORT_NORMS.get(key, PORT_NORMS.get(s.lower().strip(), s.strip()))

def norm_lane(lane_raw):
    lane = re.sub(r'\s*\(.*','',str(lane_raw)).strip()
    m = re.match(r'^(.+?)\s*[-–]\s*(.+)$', lane)
    if m: return f"{normalize_port(m.group(1))}-{normalize_port(m.group(2))}"
    return lane

def parse_num(v):
    if v is None: return None
    s = str(v).strip().replace(',','')
    if not s or s=='-' or s.lower()=='n/a': return None
    if '+' in s:
        try: return sum(float(x) for x in s.split('+'))
        except: return None
    try: return float(s)
    except: return None

def get_dest_factory(pod, lane='', consignee=''):
    p = str(pod or '').lower().strip()
    l = str(lane or '').lower()
    c = str(consignee or '').lower()
    if 'izmit' in p or 'izmit' in l or 'istanbul' in p: return 'AITK'
    if 'manzanillo' in p or 'manzanillo' in l: return 'AIMX'
    if any(x in p for x in ['hou','houston','mcallen','detroit','savanna','savannah','winchester','highland']): return 'AIUS'
    if 'aius' in c: return 'AIUS'
    if p in ['sha','shanghai'] or 'aish' in c: return 'AISH'
    if 'laem' in p or 'chabang' in p or 'rayong' in p: return 'AITH'
    if 'hamburg' in p or 'hamburg' in l or 'koper' in p or 'koper' in l: return 'AISK'
    return None

wb = openpyxl.load_workbook(EXCEL, data_only=True)
records = []

for sheet_name, period in PERIOD_MAP.items():
    if sheet_name not in wb.sheetnames: continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    sort_idx = ALL_PERIODS.index(period)
    for row in rows:
        if len(row) < 6: continue
        lane_raw = str(row[0] or '').strip()
        if not lane_raw or lane_raw.lower() in ('lane','航线','pol-pod',''): continue
        if len(lane_raw) > 60: continue
        consignor  = str(row[1] or '').strip()
        consignee  = str(row[2] or '').strip()
        pol_raw    = str(row[3] or '').strip()
        pod_raw    = str(row[4] or '').strip()
        incoterm   = str(row[5] or '').strip()
        demand     = parse_num(row[6])  if len(row)>6  else None
        booking    = parse_num(row[7])  if len(row)>7  else None
        doc        = parse_num(row[8])  if len(row)>8  else None
        thc        = parse_num(row[9])  if len(row)>9  else None
        loading    = parse_num(row[10]) if len(row)>10 else None
        customs    = parse_num(row[11]) if len(row)>11 else None
        bl         = parse_num(row[12]) if len(row)>12 else None
        local_sub  = parse_num(row[13]) if len(row)>13 else None
        dest_sub   = parse_num(row[14]) if len(row)>14 else None
        sea_fr     = parse_num(row[15]) if len(row)>15 else None
        transit    = str(row[16] or '').strip() if len(row)>16 else ''
        total      = parse_num(row[17]) if len(row)>17 else None
        forwarder  = str(row[18] or '').strip() if len(row)>18 else ''

        if pod_raw.lower() in ('pod','目的港',''): continue
        if total is None and sea_fr is None: continue

        dest = get_dest_factory(pod_raw, lane_raw, consignee)
        if not dest: continue

        lane_norm = norm_lane(lane_raw)
        records.append({
            'sheet':sheet_name,'period':period,'sort':sort_idx,
            'dest_factory':dest,'lane':lane_norm,'lane_raw':lane_raw,
            'consignor':consignor,'consignee':consignee,
            'pol':pol_raw,'pod':pod_raw,'incoterm':incoterm,
            'demand':demand,'booking':booking,'doc':doc,'thc':thc,
            'loading':loading,'customs':customs,'bl':bl,
            'local_sub':local_sub,'dest_sub':dest_sub,
            'sea_freight':sea_fr,'transit':transit,'total':total,'forwarder':forwarder,
        })

# Remove Koper and AISH Laem
records = [r for r in records if not (r['dest_factory']=='AISK' and 'koper' in r['lane'].lower())]
records = [r for r in records if not (r['dest_factory']=='AISH' and 'laem' in r['lane'].lower())]
print(f"Records: {len(records)}")

FACTORY_INFO = {
    'AITK':{'name':'AI Turkey',  'location':'土耳其 · Izmit',           'flag':'🇹🇷','color':'#3b82f6'},
    'AIMX':{'name':'AI Mexico',  'location':'墨西哥 · Manzanillo',       'flag':'🇲🇽','color':'#10b981'},
    'AIUS':{'name':'AI US',      'location':'美国 · HOU/Detroit/Savanna','flag':'🇺🇸','color':'#f59e0b'},
    'AISH':{'name':'AI Shanghai','location':'中国 · 上海（回国入境）',   'flag':'🇨🇳','color':'#f97316'},
    'AITH':{'name':'AI Thailand','location':'泰国 · Laem Chabang',       'flag':'🇹🇭','color':'#06b6d4'},
    'AISK':{'name':'AI Slovakia','location':'欧洲 · Hamburg',            'flag':'🇸🇰','color':'#8b5cf6'},
}

# Build lane trend data
tree = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
for r in records:
    if r['total']: tree[r['dest_factory']][r['lane']][r['period']].append(r['total'])

DATA = {}
FACTORY_ORDER = ['AITK','AIMX','AIUS','AISH','AITH','AISK']
for fid in FACTORY_ORDER:
    info = FACTORY_INFO[fid]
    lanes_data = {}
    for lane, period_map in tree.get(fid,{}).items():
        trend = []
        for p in ALL_PERIODS:
            vals = period_map.get(p,[])
            avg = round(statistics.mean(vals)) if vals else None
            trend.append({'period':p,'avg_total':avg,'n':len(vals)})
        curr_vals = period_map.get(CURR_PERIOD,[])
        curr_avg = round(statistics.mean(curr_vals)) if curr_vals else None
        prev_list = [t for t in trend if ALL_PERIODS.index(t['period'])<CURR_IDX and t['avg_total']]
        prev = prev_list[-1] if prev_list else None
        mom = round((curr_avg-prev['avg_total'])/prev['avg_total']*100,1) if curr_avg and prev else None
        lanes_data[lane] = {'trend':trend,'curr_avg':curr_avg,
            'prev_avg':prev['avg_total'] if prev else None,
            'prev_period':prev['period'] if prev else None,'mom':mom}
    recs = [r for r in records if r['dest_factory']==fid]
    DATA[fid] = {**info,'lanes':lanes_data,'records':recs}

with open('data_v4.json','w') as f:
    json.dump(DATA,f,ensure_ascii=False,default=str)
print("Saved data_v4.json")
